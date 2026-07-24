import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter, range_boundaries
import re
import io
import requests
from copy import copy
import os
import hashlib
from openpyxl import load_workbook
import streamlit.components.v1 as components
import base64
from datetime import datetime
import time
import threading
import queue
from concurrent.futures import ThreadPoolExecutor, as_completed
import pickle
import pyarrow as pa
import pyarrow.parquet as pq
import json
import traceback
import gc

#--- PAGE CONFIGURATION ---
st.set_page_config(
    page_title="Prime Philippines | Site Sourcing",
    layout="wide",
    initial_sidebar_state="expanded"
)

#--- HTML FRAMEWORK (UNTOUCHED BACKEND REPORT LAYOUT) ---
HTML_FRAMEWORK = """
<!DOCTYPE html>
<html>
<head>
<style type="text/css">
html, body { margin: 0; padding: 0; background-color: #ffffff; font-family: Arial, sans-serif; height: 100%; overflow-y: auto !important; overflow-x: hidden !important; }
.report-wrapper { width: 100%; overflow-y: visible !important; overflow-x: hidden !important; display: flex; justify-content: center; align-items: flex-start; padding: 0; margin: 0; min-height: 100%; }
.report-scaler { transform-origin: center top; width: 100%; display: inline-block; overflow: visible; text-align: center; }
.ritz.grid-container { height: auto; overflow: visible !important; padding: 10px; box-sizing: border-box; width: 100%; display: inline-block; text-align: left; }
.ritz .waffle a { color: inherit; }
.ritz .waffle td { padding: 2px 3px !important; vertical-align: middle; border: none !important; }
.freezebar-origin-ltr { background-color: #cecece; border: none !important; }
.column-headers-background { background-color: #cecece; text-align: center; font-size: 10pt; color: #444746; font-weight: normal; border: none !important; }
.row-headers-background { background-color: #cecece; text-align: center; font-size: 10pt; color: #444746; font-weight: normal; border: none !important; }
.ritz .waffle .s0 {border-bottom:1px SOLID #bfbfbf;border-right:1px SOLID #bfbfbf;background-color:#800000;text-align:center;font-weight:bold;color:#ffffff;font-size:10pt;white-space:nowrap;direction:ltr;padding: 4px 3px !important;}
.ritz .waffle .s1 {border-bottom:1px SOLID #bfbfbf;border-right:1px SOLID #bfbfbf;background-color:#ffffff;text-align:left;font-weight:bold;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;padding: 4px 3px !important;}
.ritz .waffle .s2 {background-color:#ffffff;text-align:left;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;border: none !important;}
.ritz .waffle .s3 {border: none !important;background-color:#ffffff;text-align:left;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;}
.ritz .waffle .s4 {border: none !important;background-color:#cecece;text-align:left;color:#000000;font-size:10pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding: 4px 3px !important;line-height: 1.4;max-width: 0;overflow: hidden;text-overflow: ellipsis;}
.ritz .waffle .s4.wrap-text {white-space:normal !important;word-wrap:break-word !important;word-break:break-word !important;overflow-wrap:break-word !important;max-width: 100% !important;overflow: visible !important;text-overflow: clip !important;height: auto !important;}
.ritz .waffle .s5 {background-color:#ffffff;text-align:left;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;border: none !important;}
.ritz .waffle .s6 {border: none !important;background-color:#ffffff;text-align:left;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;}
.ritz .waffle .s7 {border: none !important;background-color:#ffffff;text-align:left;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;}
.ritz .waffle .s8 {border: none !important;background-color:#ffffff;text-align:left;color:#ff0000;font-size:10pt;white-space:nowrap;direction:ltr;}
.ritz .waffle .s9 {border: none !important;background-color:#cecece;text-align:left;color:#000000;font-size:10pt;vertical-align:middle;white-space:nowrap;direction:ltr;padding: 4px 3px !important;line-height: 1.4;max-width: 0;overflow: hidden;text-overflow: ellipsis;}
.ritz .waffle .s9.wrap-text {white-space:normal !important;word-wrap:break-word !important;word-break:break-word !important;overflow-wrap:break-word !important;max-width: 100% !important;overflow: visible !important;text-overflow: clip !important;height: auto !important;}
.ritz .waffle .s10{background-color:#cecece;text-align:left;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;border: none !important;}
.ritz .waffle .s11{border: none !important;background-color:#ffffff;text-align:left;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;}
.ritz .waffle .s12{border: none !important;background-color:#ffffff;text-align:left;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;}
.ritz .waffle .s13{background-color:#cecece;text-align:left;font-weight:bold;color:#ff0000;font-size:10pt;white-space:nowrap;direction:ltr;border: none !important;}
.ritz .waffle .s14{background-color:#cecece;text-align:left;color:#ff0000;font-size:10pt;white-space:nowrap;direction:ltr;border: none !important;}
.ritz .waffle .s15{border: none !important;background-color:#cecece;text-align:left;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;}
.ritz .waffle .s16{border: none !important;background-color:#cecece;text-align:left;color:#ff0000;font-size:10pt;white-space:nowrap;direction:ltr;}
.ritz .waffle .s17{background-color:#cecece;text-align:left;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;border: none !important;}
.ritz .waffle .s18{border: none !important;background-color:#cecece;text-align:left;color:#ff0000;font-size:10pt;white-space:nowrap;direction:ltr;}
.ritz .waffle .s19{border: none !important;background-color:#cecece;text-align:left;color:#ff0000;font-size:10pt;white-space:nowrap;direction:ltr;}
.ritz .waffle .s20{border: none !important;background-color:#cecece;text-align:left;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;}
.ritz .waffle .s21{border: none !important;background-color:#cecece;text-align:left;color:#ff0000;font-size:10pt;white-space:nowrap;direction:ltr;}
.ritz .waffle .s22{background-color:#ffffff;text-align:left;font-weight:bold;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;border: none !important;}
.ritz .waffle .s23{border: none !important;background-color:#ffffff;text-align:left;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;}
.ritz .waffle .s24{border: none !important;background-color:#ffffff;text-align:left;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;}
.ritz .waffle .s25{border: none !important;background-color:#ffffff;text-align:left;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;}
.ritz .waffle .s26{background-color:#cecece;text-align:left;font-weight:bold;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;border: none !important;}
.ritz .waffle .s27{background-color:#cecece;text-align:left;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;border: none !important;}
.ritz .waffle .s28{background-color:#cecece;text-align:left;font-weight:bold;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;border: none !important;}
.ritz .waffle .s29{background-color:#cecece;text-align:left;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;border: none !important;}
.ritz .waffle .s30{background-color:#cecece;text-align:left;color:#000000;font-size:10pt;white-space:nowrap;direction:ltr;border: none !important;padding: 4px 3px !important;}
.ritz .waffle { border-collapse: collapse; width: 100%; }
.ritz .waffle tr { height: auto !important; }
.ritz .waffle td[class*="s4"], .ritz .waffle td[class*="s9"] { height: auto !important; min-height: 20px; }
.remarks-row { height: auto !important; }
.remarks-row td { height: auto !important; padding: 6px 3px !important; vertical-align: top !important; }
.remarks-row td.s5 { white-space: normal !important; word-wrap: break-word !important; word-break: break-word !important; overflow-wrap: break-word !important; max-width: 100% !important; overflow: visible !important; text-overflow: clip !important; height: auto !important; line-height: 1.6 !important; padding: 8px 6px !important; }
.remarks-label { white-space: nowrap !important; vertical-align: top !important; padding-top: 8px !important; }
.regulatory-header { background-color: #cecece !important; font-weight: bold !important; color: #000000 !important; }
.regulatory-label { background-color: #cecece !important; color: #ff0000 !important; }
.regulatory-value { background-color: #cecece !important; color: #000000 !important; }
</style>
<script>
document.addEventListener('DOMContentLoaded', function() {
    function scaleReportToFit() {
        const wrapper = document.querySelector('.report-wrapper');
        const scaler = document.querySelector('.report-scaler');
        const container = document.querySelector('.ritz.grid-container');
        if (!wrapper || !scaler || !container) return;
        const table = document.querySelector('.waffle');
        if (!table) return;
        const containerWidth = wrapper.parentElement ? wrapper.parentElement.clientWidth : window.innerWidth;
        const availableWidth = containerWidth - 40;
        table.style.width = 'auto';
        const naturalWidth = table.scrollWidth;
        table.style.width = '100%';
        let scale = 1;
        if (naturalWidth > availableWidth) {
            scale = (availableWidth / naturalWidth) * 0.95;
            if (scale < 0.4) scale = 0.4;
        }
        if (scale < 1) {
            scaler.style.transform = 'scale(' + scale + ')';
            scaler.style.transformOrigin = 'center top';
            scaler.style.width = (100 / scale) + '%';
            scaler.style.marginBottom = '0px';
            scaler.style.display = 'inline-block';
            scaler.style.textAlign = 'center';
        } else {
            scaler.style.transform = 'none';
            scaler.style.width = '100%';
            scaler.style.marginBottom = '0px';
            scaler.style.display = 'inline-block';
            scaler.style.textAlign = 'center';
        }
        container.style.display = 'inline-block';
        container.style.textAlign = 'left';
        container.style.margin = '0 auto';
        wrapper.style.overflowY = 'visible';
        wrapper.style.overflowX = 'hidden';
        wrapper.style.width = '100%';
        wrapper.style.display = 'flex';
        wrapper.style.justifyContent = 'center';
        wrapper.style.alignItems = 'flex-start';
        wrapper.style.minHeight = '100%';
        container.style.overflowX = 'visible';
        container.style.overflowY = 'visible';
        container.style.width = '100%';
        document.body.style.overflowY = 'auto';
        document.body.style.overflowX = 'hidden';
        document.documentElement.style.overflowY = 'auto';
        document.documentElement.style.overflowX = 'hidden';
        const tableHeight = table.scrollHeight;
        if (scale < 1) {
            scaler.style.height = (tableHeight * scale + 50) + 'px';
        } else {
            scaler.style.height = 'auto';
        }
    }
    setTimeout(scaleReportToFit, 100);
    let resizeTimer;
    window.addEventListener('resize', function() {
        clearTimeout(resizeTimer);
        resizeTimer = setTimeout(scaleReportToFit, 100);
    });
    const observer = new MutationObserver(function() { setTimeout(scaleReportToFit, 100); });
    const tableBody = document.querySelector('.waffle tbody');
    if (tableBody) { observer.observe(tableBody, { childList: true, subtree: true, characterData: true }); }
    window.addEventListener('load', function() { setTimeout(scaleReportToFit, 200); });
});
</script>
</head>
<body>
<div class="report-wrapper">
    <div class="report-scaler">
        <div class="ritz grid-container" dir="ltr">
        <table class="waffle" cellspacing="0" cellpadding="0" style="table-layout: fixed; width: 100%; border-collapse: collapse;">
        <colgroup>
        <col style="width:223px;"> <col style="width:100px;"> <col style="width:86px;"> <col style="width:100px;"> <col style="width:94px;"> <col style="width:100px;"> <col style="width:81px;"> <col style="width:15px;"> <col style="width:148px;"> <col style="width:176px;"> <col style="width:100px;"> <col style="width:100px;"> <col style="width:100px;"> <col style="width:125px;"> <col style="width:29px;">
        </colgroup>
        <tbody>
        <tr style="height: auto;"> <td class="s0" colspan="15">SITE INFORMATION REPORT</td> </tr>
        <tr style="height: auto"> <td class="s1" colspan="7">General Information</td> <td class="s1"></td> <td class="s1" colspan="7">Location</td> </tr>
        <tr style="height: auto"> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s3"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s3"></td> </tr>
        <tr style="height: auto;"> <td class="s2">Trade Area Name</td> <td class="s2"></td> <td class="s4" colspan="5">_TRADE_AREA_</td> <td class="s3"></td> <td class="s5" colspan="2">Site Name</td> <td class="s4" colspan="5">_SITE_NAME_</td> </tr>
        <tr style="height: auto;"> <td class="s2">Site Name:</td> <td class="s2"></td> <td class="s4" colspan="5">_SITE_NAME_</td> <td class="s3"></td> <td class="s5" colspan="2">Unit #, Bldg/St # and St Name</td> <td class="s4" colspan="5">_UNIT_BLDG_ST_NAME_</td> </tr>
        <tr style="height: auto;"> <td class="s2">Site Number:</td> <td class="s2"></td> <td class="s4" colspan="5">_SITE_NO_</td> <td class="s3"></td> <td class="s5" colspan="2">Barangay/District Name</td> <td class="s4" colspan="5">_BARANGAY_DISTRICT_NAME_</td> </tr>
        <tr style="height: auto;"> <td class="s2">Date Started</td> <td class="s2"></td> <td class="s4" colspan="5">_TIMESTAMP_</td> <td class="s3"></td> <td class="s5" colspan="2">City/Municipality</td> <td class="s4" colspan="5">_CITY_MUNICIPALITY_</td> </tr>
        <tr style="height: auto;"> <td class="s5" colspan="2">Date Report Submitted</td> <td class="s4" colspan="5">_DATE_OF_REPORT_</td> <td class="s3"></td> <td class="s5" colspan="2">Region</td> <td class="s4" colspan="5">_REGION_</td> </tr>
        <tr style="height: auto;"> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s3"></td> <td class="s5" colspan="2">Postal Code</td> <td class="s4" colspan="5">_POSTAL_CODE_</td> </tr>
        <tr style="height: 9px"> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s3"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s7"></td> </tr>
        <tr style="height: 19px"> <td class="s1" colspan="7">Terms</td> <td class="s3"></td> <td class="s1" colspan="7">Rates</td> </tr>
        <tr style="height: 19px"> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s3"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s3"></td> </tr>
        <tr style="height: auto;"> <td class="s2">Site Availability Date</td> <td class="s2"></td> <td class="s4" colspan="5">_SITE_AVAILABILITY_DATE_</td> <td class="s3"></td> <td class="s8" colspan="2">Monthly Rental Rate (Php)</td> <td class="s4" colspan="5">_MONTHLY_RENTAL_RATE_</td> </tr>
        <tr style="height: auto;"> <td class="s2">COL Start Date</td> <td class="s2"></td> <td class="s4" colspan="5">_COL_START_DATE_</td> <td class="s3"></td> <td class="s8" colspan="2">Percentage Rent</td> <td class="s4" colspan="5">_PERCENTAGE_RENT_</td> </tr>
        <tr style="height: auto;"> <td class="s2">COL End Date</td> <td class="s2"></td> <td class="s4" colspan="5">_COL_END_DATE_</td> <td class="s3"></td> <td class="s8" colspan="2">Minimum Guaranteed Rent</td> <td class="s4" colspan="5">_MINIMUM_GUARANTEED_RENT_</td> </tr>
        <tr style="height: auto;"> <td class="s2">Lease Terms</td> <td class="s2"></td> <td class="s4" colspan="5">_LEASE_TERMS_</td> <td class="s3"></td> <td class="s8" colspan="2">Annual Escalation Rate (%)</td> <td class="s4" colspan="5">_ESCALATION_</td> </tr>
        <tr style="height: 19px"> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s3"></td> <td class="s8" colspan="2">Advance Rental (Php)</td> <td class="s4" colspan="5">_ADVANCE_RENTAL_</td> </tr>
        <tr style="height: 19px"> <td class="s1" colspan="7">Technical Info</td> <td class="s3"></td> <td class="s8" colspan="2">Security Deposit Amount (Php)</td> <td class="s4" colspan="5">_SECURITY_DEPOSIT_</td> </tr>
        <tr style="height: 19px"> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s3"></td> <td class="s8" colspan="2">CUSA Dues</td> <td class="s4" colspan="5">_CUSA_</td> </tr>
        <tr style="height: auto;"> <td class="s5" colspan="2">Lot/Floor Area (in sqm)</td> <td class="s4" colspan="5">_LOT_FLOOR_AREA_SQM_</td> <td class="s3"></td> <td class="s8" colspan="2">Estimated Revenue Per Mo.</td> <td class="s4" colspan="5">_ESTIMATED_REVENUE_PER_MO_</td> </tr>
        <tr style="height: auto;"> <td class="s2">Frontage (in m)</td> <td class="s2"></td> <td class="s4" colspan="5">_FRONTAGE_</td> <td class="s3"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s7"></td> </tr>
        <tr style="height: auto;"> <td class="s2">Depth (in m)</td> <td class="s2"></td> <td class="s4" colspan="5">_DEPTH_IN_M_</td> <td class="s3"></td> <td class="s1" colspan="7">Provisions</td> </tr>
        <tr style="height: auto;"> <td class="s5" colspan="2">Floor to Slab Height (in m) - if Bldg</td> <td class="s4" colspan="5">_FLOOR_TO_SLAB_HEIGHT_IN_M_IF_BLDG_</td> <td class="s3"></td> <td class="s2" colspan="7"></td> </tr>
        <tr style="height: auto;"> <td class="s5" colspan="2">No. of Storeys (If Bldg Lessee)</td> <td class="s4" colspan="5">_NO_OF_STOREYS_</td> <td class="s3"></td> <td class="s5" colspan="2">Tenant is the Owner</td> <td class="s9" colspan="5">_TENANT_IS_THE_OWNER_</td> </tr>
        <tr style="height: auto;"> <td class="s5" colspan="2">Type of Structure (if Bldg Lessee)</td> <td class="s4" colspan="5">_TYPE_OF_STRUCTURE_IF_BLDG_LESSEE_</td> <td class="s3"></td> <td class="s5" colspan="2">Lease Type</td> <td class="s9" colspan="5">_LEASE_TYPE_</td> </tr>
        <tr style="height: auto;"> <td class="s2">Soil Profile</td> <td class="s2"></td> <td class="s4" colspan="5">_SOIL_PROFILE_</td> <td class="s3"></td> <td class="s5" colspan="2">Principal COL</td> <td class="s9" colspan="5">_PRINCIPAL_COL_</td> </tr>
        <tr style="height: auto;"> <td class="s2">Supply Access:</td> <td class="s2"></td> <td class="s2" colspan="5"></td> <td class="s3"></td> <td class="s5" colspan="2">Sub-Lease Provision</td> <td class="s9" colspan="5">_SUB_LEASE_PROVISION_</td> </tr>
        <tr style="height: auto;"> <td class="s2">Power</td> <td class="s10"></td> <td class="s2">Aircon</td> <td class="s10"></td> <td class="s5" colspan="2">LPG Fire Pro</td> <td class="s10"></td> <td class="s3"></td> <td class="s5" colspan="2">Pre-Term/Partial Term</td> <td class="s9" colspan="5">_PRE_TERM_PARTIAL_TERM_</td> </tr>
        <tr style="height: auto;"> <td class="s2">Water</td> <td class="s10"></td> <td class="s2">Exhaust</td> <td class="s10"></td> <td class="s5" colspan="2">Drainage TP</td> <td class="s10"></td> <td class="s3"></td> <td class="s5" colspan="2">Tripartite Agreement</td> <td class="s9" colspan="5">_TRIPARTITE_AGREEMENT_</td> </tr>
        <tr style="height: 9px;"> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s3"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s7"></td> </tr>
        <tr style="height: 19px;"> <td class="s1" colspan="7">Lessor and Tenant Details</td> <td class="s3"></td> <td class="s1" colspan="7">If with Sub-Lessor/ Sub-Lessee</td> </tr>
        <tr style="height: 9px;"> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s3"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s3"></td> </tr>
        <tr style="height: auto;"> <td class="s2">Name of Lessor</td> <td class="s2"></td> <td class="s4" colspan="5">_LESSOR_</td> <td class="s3"></td> <td class="s5" colspan="2">Name of Sub-Lessor</td> <td class="s9" colspan="5">_NAME_OF_SUBLESSOR_</td> </tr>
        <tr style="height: auto;"> <td class="s2">Contact No.</td> <td class="s2"></td> <td class="s4" colspan="5">_LESSOR_CONTACT_NO_</td> <td class="s3"></td> <td class="s5" colspan="2">Contact No.</td> <td class="s9" colspan="5">_SUBLESSOR_CONTACT_NO_</td> </tr>
        <tr style="height: auto;"> <td class="s2">E-mail Address</td> <td class="s2"></td> <td class="s4" colspan="5">_LESSOR_EMAIL_ADDRESS_</td> <td class="s3"></td> <td class="s5" colspan="2">E-mail Address</td> <td class="s9" colspan="5">_SUBLESSOR_EMAIL_ADDRESS_</td> </tr>
        <tr style="height: auto;"> <td class="s2">Type of Ownership</td> <td class="s2"></td> <td class="s4" colspan="5">_LESSOR_TYPE_OF_OWNERSHIP_</td> <td class="s3"></td> <td class="s5" colspan="2">Type of Ownership</td> <td class="s9" colspan="5">_SUBLESSOR_TYPE_OF_OWNERSHIP_</td> </tr>
        <tr style="height: auto;"> <td class="s2">Company Name</td> <td class="s2"></td> <td class="s4" colspan="5">_LESSOR_COMPANY_NAME_</td> <td class="s3"></td> <td class="s5" colspan="2">Company Name</td> <td class="s9" colspan="5">_SUBLESSOR_COMPANY_NAME_</td> </tr>
        <tr style="height: auto;"> <td class="s5" colspan="2">Developer Account Name</td> <td class="s4" colspan="5">_LESSOR_DEVELOPER_ACCOUNT_NAME_</td> <td class="s3"></td> <td class="s5" colspan="2">Developer Account Name</td> <td class="s9" colspan="5">_SUBLESSOR_DEVELOPER_ACCOUNT_NAME_</td> </tr>
        <tr style="height: auto;"> <td class="s2">Business Address</td> <td class="s2"></td> <td class="s4" colspan="5">_LESSOR_BUSINESS_ADDRESS_</td> <td class="s3"></td> <td class="s5" colspan="2">Business Address</td> <td class="s9" colspan="5">_SUBLESSOR_BUSINESS_ADDRESS_</td> </tr>
        <tr style="height: auto;"> <td class="s5" colspan="2">Name of Authorized Representative</td> <td class="s4" colspan="5">_LESSOR_AUTHORIZED_REPRESENTATIVE_</td> <td class="s3"></td> <td class="s5" colspan="2">Name of Authorized Representative</td> <td class="s9" colspan="5">_SUBLESSOR_NAME_OF_AUTHORIZED_REPRESENTATIVE_</td> </tr>
        <tr style="height: auto;"> <td class="s5" colspan="2">Residence Address of Authorized Representative</td> <td class="s4" colspan="5">_LESSOR_RESIDENCE_ADDRESS_OF_AUTHORIZED_REPRESENTATIVE_</td> <td class="s3"></td> <td class="s5" colspan="2">Residence Address of Authorized Representative</td> <td class="s9" colspan="5">_SUBLESSOR_RESIDENCE_ADDRESS_OF_AUTHORIZED_REPRESENTATIVE_</td> </tr>
        <tr style="height: auto;"> <td class="s2">Contact No.</td> <td class="s2"></td> <td class="s4" colspan="5">_LESSOR_AUTHORIZED_REP_CONTACT_NO_</td> <td class="s3"></td> <td class="s5" colspan="2">Contact No.</td> <td class="s9" colspan="5">_SUBLESSOR_CONTACT_NO_</td> </tr>
        <tr style="height: auto;"> <td class="s2">E-mail Address</td> <td class="s2"></td> <td class="s4" colspan="5">_LESSOR_AUTHORIZED_REP_EMAIL_</td> <td class="s3"></td> <td class="s5" colspan="2">E-mail Address</td> <td class="s9" colspan="5">_SUBLESSOR_EMAIL_ADDRESS_</td> </tr>
        <tr style="height: auto;"> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s3"></td> <td class="s2"></td> <td class="s2"></td> <td class="s2"></td> <td class="s3" colspan="5"></td> </tr>
        <tr style="height: auto;"> <td class="s2">Name of Lessee</td> <td class="s2"></td> <td class="s4" colspan="5">_NAME_OF_LESSEE_</td> <td class="s3"></td> <td class="s5" colspan="2">Name of Sub-Lessee</td> <td class="s9" colspan="5">_NAME_OF_SUB_LESSEE_</td> </tr>
        <tr style="height: auto;"> <td class="s2">Position</td> <td class="s2"></td> <td class="s4" colspan="5">_LESSEE_POSITION_</td> <td class="s3"></td> <td class="s5" colspan="2">Position</td> <td class="s9" colspan="5">_SUB_LESSEE_POSITION_</td> </tr>
        <tr style="height: auto;"> <td class="s2">Contact No.</td> <td class="s2"></td> <td class="s4" colspan="5">_LESSEE_CONTACT_NO_</td> <td class="s3"></td> <td class="s5" colspan="2">Contact No.</td> <td class="s9" colspan="5">_SUB_LESSEE_CONTACT_NO_</td> </tr>
        <tr style="height: auto;"> <td class="s2">E-mail Address</td> <td class="s2"></td> <td class="s4" colspan="5">_LESSEE_EMAIL_ADDRESS_</td> <td class="s3"></td> <td class="s5" colspan="2">E-mail Address</td> <td class="s9" colspan="5">_SUB_LESSEE_EMAIL_ADDRESS_</td> </tr>
        <tr style="height: auto;"> <td class="s5" colspan="2">Name of Authorized Representative</td> <td class="s4" colspan="5">_LESSEE_NAME_OF_AUTHORIZED_REPRESENTATIVE_</td> <td class="s3"></td> <td class="s5" colspan="2">Name of Authorized Representative</td> <td class="s9" colspan="5">_SUB_LESSEE_NAME_OF_AUTHORIZED_REPRESENTATIVE_</td> </tr>
        <tr style="height: auto;"> <td class="s2">Business Address</td> <td class="s2"></td> <td class="s4" colspan="5">_LESSEE_BUSINESS_ADDRESS_</td> <td class="s3"></td> <td class="s5" colspan="2">Business Address</td> <td class="s9" colspan="5">_SUB_LESSEE_BUSINESS_ADDRESS_</td> </tr>
        <tr style="height: 9px;"> <td class="s11"></td> <td class="s11"></td> <td class="s11"></td> <td class="s11"></td> <td class="s11"></td> <td class="s11"></td> <td class="s11"></td> <td class="s12"></td> <td class="s11"></td> <td class="s11"></td> <td class="s11"></td> <td class="s11"></td> <td class="s11"></td> <td class="s11"></td> <td class="s12"></td> </tr>
        
        <!-- REGULATORY SECTION -->
        <tr style="height: 19px;"> <td class="s13" colspan="15" style="border: 1px solid #cccccc !important;">Regulatory</td> </tr>
        <tr style="height: auto;"> 
            <td class="s14" style="border: 1px solid #cccccc !important;">Setback Requirement</td> 
            <td class="s17" colspan="3" style="border: 1px solid #cccccc !important;">_SETBACK_REQUIREMENT_</td> 
            <td class="s14" colspan="2" style="border: 1px solid #cccccc !important;">Perm Traffic Re-Routing</td> 
            <td class="s17" colspan="3" style="border: 1px solid #cccccc !important;">_PERM_TRAFFIC_RE_ROUTING_</td> 
            <td class="s14" colspan="2" style="border: 1px solid #cccccc !important;">Future Development</td> 
            <td class="s17" colspan="4" style="border: 1px solid #cccccc !important;">_FUTURE_DEVELOPMENT_</td> 
        </tr>
        <tr style="height: auto;"> 
            <td class="s14" style="border: 1px solid #cccccc !important;">Road Widening</td> 
            <td class="s17" colspan="3" style="border: 1px solid #cccccc !important;">_ROAD_WIDENING_</td> 
            <td class="s14" colspan="2" style="border: 1px solid #cccccc !important;">Perm Road Closure</td> 
            <td class="s17" colspan="3" style="border: 1px solid #cccccc !important;">_PERM_ROAD_CLOSURE_</td> 
            <td class="s14" colspan="2" style="border: 1px solid #cccccc !important;">Zoning Clearance</td> 
            <td class="s17" colspan="4" style="border: 1px solid #cccccc !important;">_ZONING_CLEARANCE_</td> 
        </tr>
        <tr style="height: auto;"> 
            <td class="s14" style="border: 1px solid #cccccc !important;">Pedestrian Overpass</td> 
            <td class="s17" colspan="3" style="border: 1px solid #cccccc !important;">_PEDESTRIAN_OVERPASS_</td> 
            <td class="s14" colspan="2" style="border: 1px solid #cccccc !important;">Infrastructure Programs</td> 
            <td class="s17" colspan="3" style="border: 1px solid #cccccc !important;">_INFRASTRUCTURE_PROGRAMS_</td> 
            <td class="s14" colspan="2" style="border: 1px solid #cccccc !important;">Gas Station</td> 
            <td class="s17" colspan="4" style="border: 1px solid #cccccc !important;">_GAS_STATION_</td> 
        </tr>
        <tr style="height: 9px;"> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s3"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s6"></td> <td class="s7"></td> </tr>
        
        <!-- SITE ACQUIRABILITY SECTION -->
        <tr style="height: 19px;"> 
            <td class="s22" colspan="3" style="border-bottom: 1px solid #000000 !important; padding-bottom: 4px !important;">Site Acquirability:</td>
            <td class="s3" colspan="12"></td>
        </tr>
        <tr style="height: auto;"> 
            <td class="s2" style="border: 1px solid #cccccc !important;">Confidence Level</td> 
            <td class="s17" colspan="6" style="border: 1px solid #cccccc !important;">_CONFIDENCE_LEVEL_</td> 
            <td class="s3"></td> 
            <td class="s22" colspan="2" style="border: 1px solid #cccccc !important;">Site Availability</td> 
            <td class="s17" colspan="5" style="border: 1px solid #cccccc !important;">_SITE_AVAILABILITY_CLASS_</td> 
        </tr>
        <tr class="remarks-row" style="height: 300px !important;"> 
            <td class="s6 remarks-label" style="white-space: nowrap; vertical-align: top; padding-top: 8px; border: 1px solid #cccccc !important;">Other Remarks:</td> 
            <td class="s5" colspan="6" style="white-space: normal; word-wrap: break-word; overflow-wrap: break-word; max-width: 100%; overflow: visible; text-overflow: clip; height: 300px !important; line-height: 1.6; padding: 8px 6px; border: 1px solid #cccccc !important; vertical-align: top;">_REMARKS_</td> 
            <td class="s3"></td> 
            <td class="s6 remarks-label" style="white-space: normal; vertical-align: top; padding-top: 8px; border: 1px solid #cccccc !important;" colspan="2">Site Availability<br>Remarks:</td> 
            <td class="s5" colspan="5" style="white-space: normal; word-wrap: break-word; overflow-wrap: break-word; max-width: 100%; overflow: visible; text-overflow: clip; height: 300px !important; line-height: 1.6; padding: 8px 6px; border: 1px solid #cccccc !important; vertical-align: top;">_SITE_AVAILABILITY_REMARKS_</td> 
        </tr>
        </tbody>
        </table>
        </div>
    </div>
</div>
</body>
</html>
"""

#--- FULL SCREEN LOADING OVERLAY ---
def get_loading_overlay_html(message="Loading data..."):
    return f"""
    <div id="loading-overlay" style="
        position: fixed; top: 0; left: 0; width: 100%; height: 100%;
        background: rgba(248, 247, 245, 0.98); /* Off-White */
        z-index: 999999; 
        display: flex; flex-direction: column;
        justify-content: center; align-items: center;
        font-family: 'Montserrat', 'Roboto', 'Segoe UI', sans-serif;
    ">
        <div style="
            width: 45px; height: 45px;
            border: 4px solid #0f1f38; /* Navy */
            border-top: 4px solid #c8a658; /* Gold */
            border-radius: 50%;
            animation: spin 1s linear infinite;
            margin-bottom: 20px;
        "></div>
        <div style="font-size: 1.1rem; color: #0f1f38; font-weight: 600; letter-spacing: 2px; text-transform: uppercase;">
            {message}
        </div>
    </div>
    <style>
        @keyframes spin {{ 
            0% {{ transform: rotate(0deg); }} 
            100% {{ transform: rotate(360deg); }} 
        }}
    </style>
    """

#--- GLOBAL STYLES (DASHBOARD UX OVERHAUL) ---
st.markdown("""
<style >
@import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@300;400;500;600;700&family=Playfair+Display:ital,wght@0,400;0,600;0,700;1,400;1,600&display=swap');

/* Global Font Settings */
* { font-family: 'Montserrat', 'Roboto', sans-serif !important; }
h1, h2, h3, h4, h5, h6 { font-family: 'Playfair Display', serif !important; color: #0f1f38 !important; }

/* Backgrounds & Base Colors */
html, body { overflow-y: auto !important; overflow-x: hidden !important; height: 100% !important; margin: 0px !important; padding: 0px !important; background-color: #f8f7f5 !important; }
.stApp, .appview-container, .main, [data-testid="stAppViewContainer"], [data-testid="stMain"] { background-color: #f8f7f5 !important; color: #0f1f38 !important; }

/* Streamlit Header Hiding */
header[data-testid="stHeader"], [data-testid="stHeader"], .stApp > header { display: none !important; }

/* Clean UI Overrides */
div[data-testid="stVerticalBlock"] > div:has(style), div[data-testid="stVerticalBlock"] > div:empty { display: none !important; }
iframe[title="streamlit_components.components.html"] { height: 1200px !important; max-height: none !important; border: none !important; width: 100% !important; overflow: hidden !important; background: #ffffff !important; box-shadow: 0 4px 15px rgba(15, 31, 56, 0.05); }

/* Tabs UX - Modernized */
button[data-baseweb="tab"] { font-size: 0.9rem !important; text-transform: uppercase !important; letter-spacing: 1.5px !important; color: #0f1f38 !important; font-weight: 600 !important; background: transparent !important; border: none !important; border-bottom: 2px solid transparent !important; padding: 1rem 1.5rem !important; transition: all 0.2s ease !important; }
button[data-baseweb="tab"][aria-selected="true"] { border-bottom: 3px solid #c8a658 !important; color: #c8a658 !important; }
button[data-baseweb="tab"]:hover { background-color: rgba(200, 166, 88, 0.05) !important; color: #c8a658 !important; }

/* Sidebar UX */
[data-testid="stSidebar"] { background-color: #0f1f38 !important; border-right: none !important; box-shadow: 4px 0 15px rgba(0,0,0,0.1); padding-top: 2rem !important;}
[data-testid="stSidebar"] * { color: #ffffff !important; }
[data-testid="stSidebar"] select { color: #0f1f38 !important; } /* Fix select text in sidebar */
[data-testid="stSidebar"] .stSelectbox label { text-transform: uppercase; font-size: 0.75rem; letter-spacing: 1px; color: #c8a658 !important; margin-bottom: 5px; }

/* Inputs and Selects (Main) */
.stSelectbox > div > div { background-color: #ffffff !important; border: 1px solid rgba(15,31,56,0.2) !important; border-radius: 0px !important; min-height: 38px !important; color: #0f1f38 !important; transition: border 0.3s ease; }
.stSelectbox > div > div:hover, .stSelectbox > div > div:focus-within { border: 1px solid #c8a658 !important; box-shadow: none !important; }
.stTextInput > div > div > input { border-radius: 0px !important; border: 1px solid rgba(15,31,56,0.2) !important; background-color: #ffffff !important; color: #0f1f38 !important; padding: 12px !important; transition: border 0.3s ease;}
.stTextInput > div > div > input:focus { border: 1px solid #c8a658 !important; box-shadow: none !important;}

/* Metric Cards */
[data-testid="stMetricValue"] { font-size: 1.5rem !important; font-weight: 700 !important; color: #c8a658 !important; font-family: 'Playfair Display', serif !important;}
[data-testid="stMetricLabel"] { font-size: 0.75rem !important; text-transform: uppercase !important; letter-spacing: 1px !important; color: #0f1f38 !important; font-weight: 600 !important; opacity: 0.7; }
div[data-testid="stMetric"] { background: #ffffff; padding: 1rem 1.5rem; border-left: 4px solid #0f1f38; box-shadow: 0 4px 15px rgba(15, 31, 56, 0.05); }

/* Buttons */
.stButton > button, .stDownloadButton > button { background-color: #c8a658 !important; color: #0f1f38 !important; border: 1px solid transparent !important; border-radius: 0px !important; padding: 0.5rem 1rem !important; font-size: 0.85rem !important; font-weight: 600 !important; text-transform: uppercase !important; letter-spacing: 1px !important; min-height: 40px !important; width: 100% !important; box-shadow: 0 4px 10px rgba(200, 166, 88, 0.2) !important; transition: all 0.2s ease-in-out !important;}
.stButton > button:hover, .stDownloadButton > button:hover { background-color: transparent !important; color: #c8a658 !important; border: 1px solid #c8a658 !important; box-shadow: none !important; transform: translateY(-1px); }

/* Sidebar Button overrides */
[data-testid="stSidebar"] .stButton > button, [data-testid="stSidebar"] .stDownloadButton > button { background-color: #c8a658 !important; color: #0f1f38 !important; border: none !important;}
[data-testid="stSidebar"] .stButton > button:hover, [data-testid="stSidebar"] .stDownloadButton > button:hover { background-color: #ffffff !important; color: #0f1f38 !important; }

/* Login Box wrapper */
.login-wrapper { max-width: 400px; margin: 10vh auto; background: #ffffff; padding: 3rem 2rem; box-shadow: 0 10px 30px rgba(15, 31, 56, 0.1); border-top: 4px solid #c8a658; text-align: center; }

._profilePreview_gzau3_63, ._link_gzau3_10, [class*='_profilePreview'], [class*='_link_gzau3'], a[href*='share.streamlit.io'], a[href*='streamlit.io'], img[src*='avatar'], [class*='avatar'], #MainMenu, button[title="View source"], .stAppDeployButton, div[data-testid="stStatusWidget"] { display: none !important; visibility: hidden !important; opacity: 0 !important; height: 0 !important; width: 0 !important; pointer-events: none !important; }
</style>
""", unsafe_allow_html=True)

#--- SECURITY OBSERVERS ---
def deploy_workspace_security_protocols():
    injected_js = """
    <script>
    (function() {
        const restrictedUrls = ["https://share.streamlit.io/user/pyscriptcli", "https://streamlit.io/cloud"];
        function checkAndBlockUrl(url) {
            if (!url) return false;
            const shouldBlock = restrictedUrls.some(blockedUrl => url.toLowerCase().trim().includes(blockedUrl.toLowerCase().trim()));
            if (shouldBlock) {
                window.stop();
                window.top ? window.top.location.href = window.location.origin : window.location.href = window.location.origin;
                return true;
            }
            return false;
        }
        document.addEventListener('click', function(e) {
            const target = e.target.closest('a');
            if (target && target.href) { if (checkAndBlockUrl(target.href)) { e.preventDefault(); e.stopPropagation(); } }
        }, true);
        const originalAssign = window.location.assign;
        window.location.assign = function(url) { if (!checkAndBlockUrl(url)) { originalAssign.apply(this, arguments); } };
        const originalReplace = window.location.replace;
        window.location.replace = function(url) { if (!checkAndBlockUrl(url)) { originalReplace.apply(this, arguments); } };
        function purgeTargetElements() {
            const targetSelectors = ["._profilePreview_gzau3_63", "._link_gzau3_10", "[class*='_profilePreview']", "[class*='_link_gzau3']", "a[href*='share.streamlit.io']", "a[href*='streamlit.io']", "img[src*='avatar']", "[class*='avatar']"];
            targetSelectors.forEach(selector => {
                document.querySelectorAll(selector).forEach(el => el.style.setProperty('display', 'none', 'important'));
                if (window.top && window.top.document) { try { window.top.document.querySelectorAll(selector).forEach(el => el.style.setProperty('display', 'none', 'important')); } catch(err) {} }
            });
        }
        purgeTargetElements();
        const layoutObserver = new MutationObserver(function() { purgeTargetElements(); });
        if (document.body) layoutObserver.observe(document.body, { childList: true, subtree: true });
        if (window.top && window.top.document && window.top.document.body) { try { layoutObserver.observe(window.top.document.body, { childList: true, subtree: true }); } catch(e) {} }
        setInterval(function() { purgeTargetElements(); try { checkAndBlockUrl(window.location.href); if (window.top && window.top !== window) { checkAndBlockUrl(window.top.location.href); } } catch(e) {} }, 1000);
    })();
    </script>
    """
    components.html(injected_js, height=0, width=0)

deploy_workspace_security_protocols()

#--- HARDCODED USERS ---
HARDCODED_USERS = {
    "regis": {"password": "trs.jfc", "permissions": {"view_sir": True, "export_sir": True}, "is_admin": False},
    "trs.aims": {"password": "trs.jfc", "permissions": {"view_sir": True, "export_sir": True}, "is_admin": False},
    "jfc": {"password": "trs.prime", "permissions": {"view_sir": True, "export_sir": True}, "is_admin": False},
    "admin": {"password": "@47t00M!!", "permissions": {"view_sir": True, "export_sir": True}, "is_admin": True}
}

#--- SESSION STATE ---
if 'authenticated' not in st.session_state: st.session_state.authenticated = False
if 'username' not in st.session_state: st.session_state.username = ""
if 'role' not in st.session_state: st.session_state.role = "member"
if 'data_loaded' not in st.session_state: st.session_state.data_loaded = False
if 'df' not in st.session_state: st.session_state.df = None
if 'placeholders' not in st.session_state: st.session_state.placeholders = None
if 'template_bytes_raw' not in st.session_state: st.session_state.template_bytes_raw = None
if 'media_data_list' not in st.session_state: st.session_state.media_data_list = None
if 'cache_version' not in st.session_state: st.session_state.cache_version = 0
if 'audit_log' not in st.session_state: st.session_state.audit_log = []
if 'refresh_log' not in st.session_state: st.session_state.refresh_log = []
if 'cached_reports' not in st.session_state: st.session_state.cached_reports = {}
if 'cache_timestamp' not in st.session_state: st.session_state.cache_timestamp = None
if 'last_refresh_time' not in st.session_state: st.session_state.last_refresh_time = None
if 'data_hash' not in st.session_state: st.session_state.data_hash = None

#--- LOGIN FUNCTION ---
def authenticate(username, password):
    if username in HARDCODED_USERS and HARDCODED_USERS[username]["password"] == password:
        st.session_state.authenticated = True
        st.session_state.username = username
        st.session_state.role = "admin" if HARDCODED_USERS[username]["is_admin"] else "member"
        st.session_state.audit_log.append({"user": username, "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
        return True
    return False

#--- CONFIGURATION ---
SOURCE_URL = "https://docs.google.com/spreadsheets/d/14nhO9u7zJRcOoux8I7l2IzwU7iQZNW9fRX6TCip47CE/export?format=xlsx"
TEMPLATE_URL = "https://docs.google.com/spreadsheets/d/1uS3xmnPi0o4c_EayQtURYDSMMPRDRGSb/export?format=xlsx"

#--- DATA FORMATTING CONFIGURATION ---
FORMAT_CONFIG = {
    'TIMESTAMP': 'date', 'DATE OF REPORT': 'date', 'SITE AVAILABILITY DATE': 'date',
    'COL START DATE': 'date', 'COL END DATE': 'date', 'MONTHLY RENTAL RATE': 'number',
    'PERCENTAGE RENT': 'number', 'MINIMUM GUARANTEED RENT': 'number', 'ESCALATION': 'number',
    'ADVANCE RENTAL': 'number', 'SECURITY DEPOSIT': 'number', 'CUSA': 'number',
    'ESTIMATED REVENUE PER MO.': 'number', 'LOT/FLOOR AREA SQM': 'number',
    'FRONTAGE': 'number', 'DEPTH (IN M)': 'number',
}

def format_value(value, format_type='text'):
    if pd.isna(value) or value is None: return ""
    if format_type == 'date':
        if isinstance(value, datetime): return value.strftime('%B %d, %Y')
        if isinstance(value, str):
            try:
                for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%d/%m/%Y']:
                    try: return datetime.strptime(value.strip(), fmt).strftime('%B %d, %Y')
                    except ValueError: continue
            except: pass
            return value
        if isinstance(value, (int, float)):
            try:
                from datetime import timedelta
                return (datetime(1899, 12, 30) + timedelta(days=float(value))).strftime('%B %d, %Y')
            except: pass
        return str(value)
    elif format_type == 'number':
        try:
            num = float(str(value).replace(',', ''))
            return f"{int(num):,}" if num.is_integer() else f"{num:,.2f}"
        except: return str(value)
    elif format_type == 'currency':
        try:
            num = float(str(value).replace(',', '').replace('₱', ''))
            return f"₱{int(num):,}" if num.is_integer() else f"₱{num:,.2f}"
        except: return str(value)
    elif format_type == 'percent':
        try: return f"{float(str(value).replace('%', ''))}%"
        except: return str(value)
    else: return str(value)

def extract_photo_url_from_cell(cell):
    if cell is None: return ""
    raw_val = cell.value if hasattr(cell, 'value') else cell
    if raw_val is None: return ""
    val_str = str(raw_val).strip()
    image_formula_match = re.search(r'IMAGE\s*\(\s*["\'](https?://[^"\']+)["\']', val_str, re.IGNORECASE)
    if image_formula_match: return image_formula_match.group(1)
    url_match = re.search(r'(https?://[^\s"\']+)', val_str)
    if url_match: return url_match.group(1)
    return val_str

def download_file(url):
    try:
        response = requests.get(url, timeout=30)
        return io.BytesIO(response.content) if response.status_code == 200 else None
    except Exception: return None

def clean_and_extract_url(cell_value):
    if cell_value is None: return ""
    val_str = str(cell_value).strip()
    formula_match = re.search(r'IMAGE\s*\(\s*["\'](https://[^"\']+)["\']', val_str, re.IGNORECASE)
    if formula_match: return formula_match.group(1)
    url_match = re.search(r'(https://[^\s"\']+)', val_str)
    if url_match: return url_match.group(1)
    return val_str

def extract_google_drive_id(clean_url):
    if not clean_url: return None
    match = re.search(r'(?:id=|/d/|/uc?.*?id=)([a-zA-Z0-9_-]{25,})', clean_url)
    return match.group(1) if match else None

def get_placeholders_optimized(template_bytes):
    try:
        wb = load_workbook(io.BytesIO(template_bytes), data_only=False)
        placeholders = {name for row in wb.active.iter_rows(values_only=True) for val in row if isinstance(val, str) for match in re.findall(r"{{(.*?)}}", val) for name in [match.split(":")[0].strip() if ":" in match else match.strip()]}
        wb.close()
        return sorted(list(placeholders))
    except Exception: return []

def sanitize_tab_name(name, existing_names):
    clean_name = re.sub(r'[/*?\[\]:]', '', str(name))[:31] or "Sheet"
    if clean_name not in existing_names:
        existing_names.add(clean_name)
        return clean_name
    counter = 1
    while f"{clean_name[:27]} ({counter})" in existing_names: counter += 1
    new_name = f"{clean_name[:27]} ({counter})"
    existing_names.add(new_name)
    return new_name

def parse_site_number(site_display_str):
    if not site_display_str: return float('inf')
    match = re.match(r"^(\d+(?:\.\d+)?)", str(site_display_str).strip())
    return float(match.group(1)) if match else float('inf')

def load_main_data_optimized(source_bytes):
    try:
        wb = load_workbook(io.BytesIO(source_bytes), data_only=True)
        ws = wb.active
        headers = [str(h).strip().upper() if h else "" for h in list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]]
        parsed_data_list = []
        for row in ws.iter_rows(min_row=2):
            row_dict = {}
            has_val = False
            for idx, cell in enumerate(row):
                if idx < len(headers) and headers[idx]:
                    cleaned_val = clean_and_extract_url(format_value(cell.value, FORMAT_CONFIG.get(headers[idx], 'text')) if cell.value is not None else "")
                    row_dict[headers[idx]] = cleaned_val
                    if cleaned_val: has_val = True
            if has_val: parsed_data_list.append(row_dict)
        wb.close()
        df = pd.DataFrame(parsed_data_list)
        df = df.loc[:, ~df.columns.str.contains('^$')]
        def create_site_display(row):
            site_no, site_name = row.get('SITE NO', ''), row.get('SITE NAME', '')
            if pd.notna(site_no) and str(site_no).strip():
                try:
                    f_val = float(str(site_no).strip())
                    return f"{int(f_val) if f_val.is_integer() else f_val} - {site_name}"
                except ValueError: return f"{site_no} - {site_name}"
            return str(site_name)
        df["SITE_DISPLAY"] = df.apply(create_site_display, axis=1)
        return df
    except Exception: return None

def load_media_data_optimized(source_bytes):
    try:
        wb = load_workbook(io.BytesIO(source_bytes), data_only=False)
        media_ws = next((wb[s] for s in wb.sheetnames if any(k in s.upper() for k in ["PHOTO", "DOC", "MEDIA"])), wb.active)
        media_data_list = []
        for row in media_ws.iter_rows():
            vals = [extract_photo_url_from_cell(cell) if cell else "" for cell in row]
            t_area, s_name = str(vals[13] if len(vals) > 13 and vals[13] else "").strip(), str(vals[15] if len(vals) > 15 and vals[15] else "").strip()
            if t_area and s_name and t_area.upper() != "TRADE AREA":
                media_data_list.append({
                    'TRADE AREA': t_area, 'SITE NAME': s_name,
                    '__DIRECT_TCT': vals[2] if len(vals) > 2 else "", '__DIRECT_LOT_PLAN': vals[3] if len(vals) > 3 else "",
                    '__DIRECT_BLDG_PLAN': vals[4] if len(vals) > 4 else "", '__DIRECT_TAX_MAP': vals[5] if len(vals) > 5 else "",
                    '__DIRECT_PHOTO_1': vals[7] if len(vals) > 7 else "", '__DIRECT_PHOTO_2': vals[8] if len(vals) > 8 else "",
                    '__DIRECT_PHOTO_3': vals[9] if len(vals) > 9 else "", '__DIRECT_PHOTO_4': vals[10] if len(vals) > 10 else "",
                    '__DIRECT_PHOTO_5': vals[11] if len(vals) > 11 else "",
                })
        wb.close()
        return media_data_list
    except Exception: return []

def download_and_process_all_data():
    with ThreadPoolExecutor(max_workers=2) as executor:
        f_src, f_tmp = executor.submit(download_file, SOURCE_URL), executor.submit(download_file, TEMPLATE_URL)
        if not (f_src.result() and f_tmp.result()): return None, None, None, [], None, None
        source_data, template_data = f_src.result().getvalue(), f_tmp.result().getvalue()
    
    with ThreadPoolExecutor(max_workers=2) as executor:
        f_main, f_media = executor.submit(load_main_data_optimized, source_data), executor.submit(load_media_data_optimized, source_data)
        if (df := f_main.result()) is None: return None, None, None, [], None, None
        media_data_list = f_media.result()
        
    placeholders = get_placeholders_optimized(template_data)
    cached_reports = {ta: generate_single_trade_area_report(ta, df, template_data, placeholders) for ta in df["TRADE AREA"].dropna().unique()}
    return df, placeholders, template_data, media_data_list, cached_reports, datetime.now()

def generate_single_trade_area_report(trade_area, df, template_bytes_raw, placeholders):
    placeholders_sorted = sorted(placeholders, key=len, reverse=True)
    ta_data = df[df["TRADE AREA"] == trade_area]
    wb = load_workbook(io.BytesIO(template_bytes_raw))
    base_sheet = wb.active
    base_sheet.title = "TEMPLATE_TO_DELETE"
    existing_tabs = set()
    for _, r_row in ta_data.iterrows():
        new_sheet = wb.copy_worksheet(base_sheet)
        new_sheet.title = sanitize_tab_name(r_row.get("SITE NAME", "Unknown"), existing_tabs)
        for row_cells in new_sheet.iter_rows():
            for cell in row_cells:
                if isinstance(cell.value, str) and "{{" in cell.value:
                    new_val = cell.value
                    for ph in placeholders_sorted:
                        target_regex = r"\{\{\s*" + re.escape(ph) + r"(\s*:.*?)?\}\}"
                        if re.search(target_regex, new_val):
                            new_val = re.sub(target_regex, str(r_row.get(ph.upper(), "") or ""), new_val)
                    cell.value = re.sub(r"\{\{.*?\}\}", "", new_val).strip()
        for row in new_sheet.iter_rows():
            if max([len(str(cell.value or '')) for cell in row]) > 45:
                new_sheet.row_dimensions[row[0].row].height = None
    if "TEMPLATE_TO_DELETE" in wb.sheetnames: wb.remove(wb["TEMPLATE_TO_DELETE"])
    for name in [n for n in wb.sheetnames if n not in existing_tabs and n != "TEMPLATE_TO_DELETE"]: wb.remove(wb[name])
    wb_buffer = io.BytesIO()
    wb.save(wb_buffer)
    return wb_buffer.getvalue()

@st.cache_data(ttl=None, show_spinner=False)
def get_cached_data(cache_version): return download_and_process_all_data()

#--- MAIN APP ROUTING ---
if not st.session_state.authenticated:
    st.markdown("""
        <div class="login-wrapper">
            <h2 style='font-size: 1.8rem; margin-bottom: 5px;'>PRIME Philippines</h2>
            <p style='color: #888; font-size: 0.9rem; text-transform: uppercase; letter-spacing: 2px; margin-bottom: 30px;'>Site Sourcing Portal</p>
        </div>
    """, unsafe_allow_html=True)
    c1, c2, c3 = st.columns([1, 1.2, 1])
    with c2:
        username = st.text_input("Username", placeholder="Enter your username", key="login_username")
        password = st.text_input("Password", placeholder="Enter your password", type="password", key="login_password")
        if st.button("Access Portal", use_container_width=True):
            if username and password:
                if authenticate(username, password):
                    st.toast('Welcome back to the portal!', icon='✅')
                    time.sleep(0.5)
                    st.rerun()
                else: st.error("Authentication failed. Please check your credentials.")
            else: st.warning("Please enter both username and password.")
    st.stop()

if not st.session_state.data_loaded:
    loading_placeholder = st.empty()
    loading_placeholder.markdown(get_loading_overlay_html(message="Synchronizing Site Data..."), unsafe_allow_html=True)
    result = get_cached_data(st.session_state.cache_version)
    
    if result and result[0] is not None:
        df, placeholders, template_bytes_raw, media_data_list, cached_reports, data_timestamp = result
        st.session_state.update(df=df, placeholders=placeholders, template_bytes_raw=template_bytes_raw, media_data_list=media_data_list, cached_reports=cached_reports, cache_timestamp=data_timestamp, last_refresh_time=datetime.now(), data_hash=hashlib.md5(pd.util.hash_pandas_object(df).values).hexdigest(), data_loaded=True)
        loading_placeholder.empty()
        st.toast('Data synchronized successfully.', icon='✅')
        st.rerun()
    else:
        loading_placeholder.empty()
        st.error("Failed to load data assets. Please verify link paths.")
        st.stop()

df, placeholders, template_bytes_raw, media_data_list, cached_reports = st.session_state.df, st.session_state.placeholders, st.session_state.template_bytes_raw, st.session_state.media_data_list, st.session_state.cached_reports

deploy_workspace_security_protocols()

#--- SIDEBAR DASHBOARD CONTROL ---
with st.sidebar:
    st.markdown("<h2 style='color: white !important; margin-bottom: 0px;'>PRIME</h2>", unsafe_allow_html=True)
    st.markdown("<p style='color: #c8a658 !important; font-size: 0.75rem; letter-spacing: 2px; text-transform: uppercase; margin-bottom: 30px;'>Site Sourcing</p>", unsafe_allow_html=True)
    
    nav_page = "Viewer"
    if st.session_state.role == "admin":
        nav_page = st.radio("Navigation", ["Viewer", "Admin Panel"], label_visibility="collapsed")
        st.markdown("<hr style='border-color: rgba(255,255,255,0.1); margin: 20px 0;'>", unsafe_allow_html=True)

    if nav_page == "Viewer":
        trade_areas = sorted(df["TRADE AREA"].dropna().unique().tolist())
        default_ta_index = trade_areas.index(df.iloc[0]["TRADE AREA"]) if not df.empty and df.iloc[0]["TRADE AREA"] in trade_areas else 0
        
        st.markdown("<p style='color: white !important; font-size: 0.9rem; font-weight: 600; margin-bottom: 10px;'>FILTER PROPERTIES</p>", unsafe_allow_html=True)
        selected_ta = st.selectbox("Select Trade Area", options=trade_areas, index=default_ta_index)
        
        if selected_ta:
            ta_df = df[df["TRADE AREA"] == selected_ta]
            def is_whole_number_site(val):
                try: return float(str(val).strip()).is_integer()
                except: return False
            main_sites_df = ta_df[ta_df["SITE NO"].apply(is_whole_number_site)]
            sites_in_ta = sorted(main_sites_df["SITE_DISPLAY"].dropna().unique().tolist(), key=parse_site_number)
            first_site_display = df.iloc[0]["SITE_DISPLAY"] if not df.empty else ""
            default_site_index = sites_in_ta.index(first_site_display) if selected_ta == df.iloc[0]["TRADE AREA"] and first_site_display in sites_in_ta else 0
        else:
            sites_in_ta, default_site_index = [], 0
            
        selected_site_display = st.selectbox("Select Site Name", options=sites_in_ta, index=default_site_index)
        
        st.markdown("<hr style='border-color: rgba(255,255,255,0.1); margin: 20px 0;'>", unsafe_allow_html=True)
        
        if selected_ta and HARDCODED_USERS.get(st.session_state.username, {}).get("permissions", {}).get("export_sir", False):
            report_bytes = cached_reports.get(selected_ta) or generate_single_trade_area_report(selected_ta, df, template_bytes_raw, placeholders)
            st.session_state.cached_reports[selected_ta] = report_bytes
            st.download_button(label="Export Report (XLSX)", data=report_bytes, file_name=f"{selected_ta}_SIR.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        else:
            st.button("Export Disabled", disabled=True, use_container_width=True)

#--- MAIN CONTENT AREA ---
if nav_page == "Admin Panel" and st.session_state.role == "admin":
    st.title("System Admin Panel")
    
    col1, col2, col3 = st.columns(3)
    with col1: st.metric("Cached Reports", len(st.session_state.cached_reports))
    with col2: st.metric("Trade Areas", len(st.session_state.df["TRADE AREA"].dropna().unique()) if st.session_state.df is not None else 0)
    with col3: 
        age = (datetime.now() - st.session_state.cache_timestamp).total_seconds() if st.session_state.cache_timestamp else 0
        st.metric("Cache Age", f"{int(age)}s" if age < 60 else f"{int(age/60)}m")
        
    st.divider()
    c1, c2 = st.columns(2)
    with c1:
        if st.button("Refresh All Data", use_container_width=True):
            st.cache_data.clear()
            st.session_state.update(cache_version=st.session_state.cache_version + 1, data_loaded=False, cached_reports={})
            st.session_state.refresh_log.append({"timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "user": st.session_state.username, "action": "Full data refresh"})
            st.toast("Refresh initiated...", icon="🔄")
            time.sleep(1)
            st.rerun()
    with c2:
        if st.button("Force Garbage Collection", use_container_width=True):
            gc.collect()
            st.toast("Memory cleaned!", icon="🧹")
            
    with st.expander("Audit Log"):
        if st.session_state.audit_log:
            st.dataframe(pd.DataFrame(st.session_state.audit_log).sort_values("timestamp", ascending=False), use_container_width=True)
        else: st.write("No audit records yet.")

elif nav_page == "Viewer":
    if selected_ta and selected_site_display:
        site_data = df[df["SITE_DISPLAY"] == selected_site_display]
        site_row_data = site_data.iloc[0] if not site_data.empty else None
        
        # Dashboard Header
        if site_row_data is not None:
            clean_name = str(site_row_data.get("SITE NAME", "Unknown Property"))
            st.markdown(f"<h1 style='margin-bottom: 5px; font-size: 2.2rem;'>{clean_name}</h1>", unsafe_allow_html=True)
            st.markdown(f"<p style='color: #888; text-transform: uppercase; letter-spacing: 1.5px; font-size: 0.8rem; margin-bottom: 20px;'>{selected_ta} Trade Area</p>", unsafe_allow_html=True)
            
            # Quick Stats Metrics
            m1, m2, m3, m4 = st.columns(4)
            area_val = str(site_row_data.get("LOT/FLOOR AREA SQM", "--"))
            rent_val = str(site_row_data.get("MONTHLY RENTAL RATE", "--"))
            avail_val = str(site_row_data.get("SITE AVAILABILITY DATE", "--"))
            type_val = str(site_row_data.get("LEASE TYPE", "--"))
            
            m1.metric("Lot / Floor Area", f"{area_val} sqm" if area_val != "--" and area_val else "--")
            m2.metric("Monthly Rent", rent_val if rent_val else "--")
            m3.metric("Availability", avail_val if avail_val else "--")
            m4.metric("Lease Type", type_val if type_val else "--")
            
            st.markdown("<br>", unsafe_allow_html=True)

        target_ta, target_sn = str(site_row_data["TRADE AREA"]), str(site_row_data["SITE NAME"])
        media_row_data = next((m for m in media_data_list if m['TRADE AREA'] == target_ta and m['SITE NAME'] == target_sn), site_row_data.to_dict() if site_row_data is not None else {})

        if HARDCODED_USERS.get(st.session_state.username, {}).get("permissions", {}).get("view_sir", False):
            tab_report, tab_photos, tab_docs = st.tabs(["📄 Information Report", "📷 Property Photos", "📋 Layout & Documents"])

            with tab_report:
                if site_row_data is not None:
                    try:
                        def process_val(row_data, key_string): return str(val) if not pd.isna(val := row_data.get(key_string.upper(), "")) and val is not None else ""
                        parent_site_no = parse_site_number(selected_site_display)
                        family_df = df[(df["TRADE AREA"] == target_ta) & (df["SITE NO"].apply(lambda x: int(float(str(x))) if str(x).replace('.','').isdigit() else -1) == int(parent_site_no))].sort_values(by="SITE NO", key=lambda col: col.astype(float))

                        for i in range(len(family_df)):
                            current_row = family_df.iloc[i]
                            rendered_view = HTML_FRAMEWORK
                            replacements = sorted([(f"_{k.replace(' ', '_').replace('/', '_').replace('.', '').replace('(', '').replace(')', '').replace('-', '_').upper()}_", process_val(current_row, k)) for k in FORMAT_CONFIG.keys()] + [
                                ("_TRADE_AREA_", process_val(current_row, "TRADE AREA")), ("_SITE_NAME_", process_val(current_row, "SITE NAME")), ("_SITE_NO_", process_val(current_row, "SITE NO")),
                                ("_UNIT_BLDG_ST_NAME_", process_val(current_row, "UNIT #, BLDG/ST # AND ST NAME")), ("_BARANGAY_DISTRICT_NAME_", process_val(current_row, "BARANGAY/DISTRICT NAME")),
                                ("_CITY_MUNICIPALITY_", process_val(current_row, "CITY/MUNICIPALITY")), ("_REGION_", process_val(current_row, "REGION")), ("_POSTAL_CODE_", process_val(current_row, "POSTAL CODE")),
                                ("_LEASE_TERMS_", process_val(current_row, "LEASE TERMS")), ("_FLOOR_TO_SLAB_HEIGHT_IN_M_IF_BLDG_", process_val(current_row, "FLOOR TO SLAB HEIGHT (IN M) - IF BLDG")),
                                ("_NO_OF_STOREYS_", process_val(current_row, "NO. OF STOREYS (IF BLDG LESSEE)")), ("_TYPE_OF_STRUCTURE_IF_BLDG_LESSEE_", process_val(current_row, "TYPE OF STRUCTURE (IF BLDG LESSEE)")),
                                ("_SOIL_PROFILE_", process_val(current_row, "SOIL PROFILE")), ("_TENANT_IS_THE_OWNER_", process_val(current_row, "TENANT IS THE OWNER")), ("_LEASE_TYPE_", process_val(current_row, "LEASE TYPE")),
                                ("_PRINCIPAL_COL_", process_val(current_row, "PRINCIPAL COL")), ("_SUB_LEASE_PROVISION_", process_val(current_row, "SUB-LEASE PROVISION")), ("_PRE_TERM_PARTIAL_TERM_", process_val(current_row, "PRE-TERM/PARTIAL TERM")),
                                ("_TRIPARTITE_AGREEMENT_", process_val(current_row, "TRIPARTITE AGREEMENT")), ("_LESSOR_", process_val(current_row, "LESSOR")), ("_LESSOR_CONTACT_NO_", process_val(current_row, "LESSOR CONTACT NO.")),
                                ("_LESSOR_EMAIL_ADDRESS_", process_val(current_row, "LESSOR E-MAIL ADDRESS")), ("_LESSOR_TYPE_OF_OWNERSHIP_", process_val(current_row, "LESSOR TYPE OF OWNERSHIP")),
                                ("_LESSOR_COMPANY_NAME_", process_val(current_row, "LESSOR COMPANY NAME")), ("_LESSOR_DEVELOPER_ACCOUNT_NAME_", process_val(current_row, "LESSOR DEVELOPER ACCOUNT NAME")),
                                ("_LESSOR_BUSINESS_ADDRESS_", process_val(current_row, "LESSOR BUSINESS ADDRESS")), ("_LESSOR_AUTHORIZED_REPRESENTATIVE_", process_val(current_row, "CONTACT PERSON/SOURCE")),
                                ("_LESSOR_RESIDENCE_ADDRESS_OF_AUTHORIZED_REPRESENTATIVE_", process_val(current_row, "LESSOR RESIDENCE ADDRESS OF AUTHORIZED REPRESENTATIVE")),
                                ("_LESSOR_AUTHORIZED_REP_CONTACT_NO_", process_val(current_row, "CONTACT NUMBER")), ("_LESSOR_AUTHORIZED_REP_EMAIL_", process_val(current_row, "EMAIL ADDRESS")),
                                ("_NAME_OF_LESSEE_", process_val(current_row, "NAME OF LESSEE")), ("_LESSEE_POSITION_", process_val(current_row, "LESSEE POSITION")), ("_LESSEE_CONTACT_NO_", process_val(current_row, "LESSEE CONTACT NO.")),
                                ("_LESSEE_EMAIL_ADDRESS_", process_val(current_row, "LESSEE E-MAIL ADDRESS")), ("_LESSEE_NAME_OF_AUTHORIZED_REPRESENTATIVE_", process_val(current_row, "LESSEE NAME OF AUTHORIZED REPRESENTATIVE")),
                                ("_LESSEE_BUSINESS_ADDRESS_", process_val(current_row, "LESSEE BUSINESS ADDRESS")), ("_NAME_OF_SUBLESSOR_", process_val(current_row, "NAME OF SUB-LESSOR")),
                                ("_SUBLESSOR_CONTACT_NO_", process_val(current_row, "SUB-LESSOR CONTACT NO.")), ("_SUBLESSOR_EMAIL_ADDRESS_", process_val(current_row, "SUB-LESSOR E-MAIL ADDRESS")),
                                ("_SUBLESSOR_TYPE_OF_OWNERSHIP_", process_val(current_row, "SUB-LESSOR TYPE OF OWNERSHIP")), ("_SUBLESSOR_COMPANY_NAME_", process_val(current_row, "SUB-LESSOR COMPANY NAME")),
                                ("_SUBLESSOR_DEVELOPER_ACCOUNT_NAME_", process_val(current_row, "SUB-LESSOR DEVELOPER ACCOUNT NAME")), ("_SUBLESSOR_BUSINESS_ADDRESS_", process_val(current_row, "SUB-LESSOR BUSINESS ADDRESS")),
                                ("_SUBLESSOR_NAME_OF_AUTHORIZED_REPRESENTATIVE_", process_val(current_row, "SUB-LESSOR NAME OF AUTHORIZED REPRESENTATIVE")),
                                ("_SUBLESSOR_RESIDENCE_ADDRESS_OF_AUTHORIZED_REPRESENTATIVE_", process_val(current_row, "SUB-LESSOR RESIDENCE ADDRESS OF AUTHORIZED REPRESENTATIVE")),
                                ("_NAME_OF_SUB_LESSEE_", process_row(current_row, "NAME OF SUB-LESSEE") if 'process_row' in globals() else process_val(current_row, "NAME OF SUB-LESSEE")),
                                ("_SUB_LESSEE_POSITION_", process_val(current_row, "SUB-LESSEE POSITION")), ("_SUB_LESSEE_CONTACT_NO_", process_val(current_row, "SUB-LESSEE CONTACT NO.")),
                                ("_SUB_LESSEE_EMAIL_ADDRESS_", process_val(current_row, "SUB-LESSEE E-MAIL ADDRESS")), ("_SUB_LESSEE_NAME_OF_AUTHORIZED_REPRESENTATIVE_", process_val(current_row, "SUB-LESSEE NAME OF AUTHORIZED REPRESENTATIVE")),
                                ("_SUB_LESSEE_BUSINESS_ADDRESS_", process_val(current_row, "SUB-LESSEE BUSINESS ADDRESS")), ("_SETBACK_REQUIREMENT_", process_val(current_row, "SETBACK REQUIREMENT")),
                                ("_ROAD_WIDENING_", process_val(current_row, "ROAD WIDENING")), ("_PEDESTRIAN_OVERPASS_", process_val(current_row, "PEDESTRIAN OVERPASS")),
                                ("_PERM_TRAFFIC_RE_ROUTING_", process_val(current_row, "PERM TRAFFIC RE-ROUTING")), ("_PERM_ROAD_CLOSURE_", process_val(current_row, "PERM ROAD CLOSURE")),
                                ("_INFRASTRUCTURE_PROGRAMS_", process_val(current_row, "INFRASTRUCTURE PROGRAMS")), ("_FUTURE_DEVELOPMENT_", process_val(current_row, "FUTURE DEVELOPMENT")),
                                ("_ZONING_CLEARANCE_", process_val(current_row, "ZONING CLEARANCE")), ("_GAS_STATION_", process_val(current_row, "GAS STATION")),
                                ("_CONFIDENCE_LEVEL_", process_val(current_row, "CONFIDENCE LEVEL")), ("_SITE_AVAILABILITY_CLASS_", process_val(current_row, "SITE AVAILABILITY CLASS")),
                                ("_SITE_AVAILABILITY_REMARKS_", process_val(current_row, "SITE AVAILABILITY REMARKS")), ("_REMARKS_", process_val(current_row, "REMARKS"))
                            ], key=lambda x: len(x[0]), reverse=True)
                            
                            for placeholder, value in replacements: rendered_view = rendered_view.replace(placeholder, value)
                            rendered_view = re.sub(r"_[A-Z0-9_]+_", "", rendered_view)
                            
                            st.markdown("<div style='background: white; padding: 20px; border-radius: 8px; box-shadow: 0 4px 15px rgba(15,31,56,0.05); margin-bottom: 20px;'>", unsafe_allow_html=True)
                            components.html(rendered_view, height=1600, scrolling=False)
                            st.markdown("</div>", unsafe_allow_html=True)
                    except Exception as e: st.error(f"Error compiling visual framework: {str(e)}")
                else: st.info("No data available.")
                    
            with tab_photos:
                if site_row_data is not None and media_row_data:
                    valid_photos = [(label, f"https://drive.google.com/thumbnail?sz=w800&id={fid}" if (fid := extract_google_drive_id(url)) else url, f"https://drive.google.com/uc?export=view&id={fid}" if fid else url) for label, key in {"Photo 1": "__DIRECT_PHOTO_1", "Photo 2": "__DIRECT_PHOTO_2", "Photo 3": "__DIRECT_PHOTO_3", "Photo 4": "__DIRECT_PHOTO_4", "Photo 5": "__DIRECT_PHOTO_5"}.items() if (url := media_row_data.get(key, ""))]
                    if valid_photos:
                        grid_html = '<style>.ig{display:grid;grid-template-columns:repeat(auto-fill,minmax(250px,1fr));gap:20px;padding:10px 0;}.ii{border:none;border-radius:8px;overflow:hidden;background:#fff;transition:transform .2s;aspect-ratio:4/3;display:flex;flex-direction:column;box-shadow:0 4px 15px rgba(15,31,56,0.08);}.ii:hover{transform:scale(1.03);box-shadow:0 8px 25px rgba(15,31,56,0.15);}.ii img{width:100%;height:100%;object-fit:cover;display:block;flex:1;}.ii .l{padding:12px;font-size:0.75rem;font-family:"Montserrat",sans-serif;font-weight:600;text-transform:uppercase;letter-spacing:1px;color:#0f1f38;text-align:center;border-top:1px solid rgba(15,31,56,0.05);}.ii a{text-decoration:none;color:inherit;display:flex;flex-direction:column;height:100%;}</style><div class="ig">'
                        for label, thumb_url, full_url in valid_photos: grid_html += f'<div class="ii"><a href="{full_url}" target="_blank"><img src="{thumb_url}" alt="{label}" loading="lazy"><div class="l">{label}</div></a></div>'
                        components.html(grid_html + '</div>', height=800, scrolling=True)
                    else: st.info("No photos available for this property.")

            with tab_docs:
                if site_row_data is not None and media_row_data:
                    valid_docs = [(label, f"https://drive.google.com/thumbnail?sz=w800&id={fid}" if (fid := extract_google_drive_id(url)) else url, f"https://drive.google.com/uc?export=view&id={fid}" if fid else url) for label, key in {"Title / TCT": "__DIRECT_TCT", "Lot Plan": "__DIRECT_LOT_PLAN", "Building Plan": "__DIRECT_BLDG_PLAN", "Tax Map": "__DIRECT_TAX_MAP"}.items() if (url := media_row_data.get(key, ""))]
                    if valid_docs:
                        grid_html = '<style>.ig{display:grid;grid-template-columns:repeat(auto-fill,minmax(250px,1fr));gap:20px;padding:10px 0;}.ii{border:none;border-radius:8px;overflow:hidden;background:#fff;transition:transform .2s;aspect-ratio:4/3;display:flex;flex-direction:column;box-shadow:0 4px 15px rgba(15,31,56,0.08);}.ii:hover{transform:scale(1.03);box-shadow:0 8px 25px rgba(15,31,56,0.15);}.ii img{width:100%;height:100%;object-fit:cover;display:block;flex:1;}.ii .l{padding:12px;font-size:0.75rem;font-family:"Montserrat",sans-serif;font-weight:600;text-transform:uppercase;letter-spacing:1px;color:#0f1f38;text-align:center;border-top:1px solid rgba(15,31,56,0.05);}.ii a{text-decoration:none;color:inherit;display:flex;flex-direction:column;height:100%;}</style><div class="ig">'
                        for label, thumb_url, full_url in valid_docs: grid_html += f'<div class="ii"><a href="{full_url}" target="_blank"><img src="{thumb_url}" alt="{label}" loading="lazy"><div class="l">{label}</div></a></div>'
                        components.html(grid_html + '</div>', height=800, scrolling=True)
                    else: st.info("No layout documents available for this property.")
        else:
            st.warning("You do not have permission to view site information reports.")
